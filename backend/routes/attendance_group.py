import cv2
import numpy as np
import insightface
import openpyxl
import gridfs
import os
import shutil
from pymongo import MongoClient
from flask import Blueprint, request, jsonify
from datetime import datetime
from numpy.linalg import norm
from dotenv import load_dotenv

from routes.jwt_middleware import token_required

# ---------------- Blueprint ----------------
group_attendance_bp = Blueprint(
    "group_attendance_bp",
    __name__,
    url_prefix="/api/attendance"
)

# ---------------- ENV ----------------
load_dotenv()
MONGO_URI = os.getenv("MONGO_URI")

ARC_THRESHOLD = 0.38
GROUP_PHOTOS_DIR = "group_photos"

# ---------------- DB ----------------
client = MongoClient(MONGO_URI)
db = client["AttendanceDB"]
students_col = db["students"]
fs = gridfs.GridFS(db)

# ---------------- MODEL ----------------
from utils.model_loader import get_model
model = get_model()

# ---------------- LOAD STUDENTS ----------------
names, rolls, encodings = [], [], []

for doc in students_col.find({"model": "arcface"}):
    names.append(doc["name"])
    rolls.append(str(doc["rollNo"]))
    encodings.append(np.array(doc["face_encoding"], dtype=float))

# ---------------- UTILS ----------------
def cosine_sim(a, b):
    return float(np.dot(a, b) / (norm(a) * norm(b)))

# ---------------- GROUP ATTENDANCE ----------------
@group_attendance_bp.route("/group", methods=["POST"])
@token_required
def group_attendance():
    teacher = request.teacher

    # -------- Metadata --------
    course = request.form.get("course", "COURSE")
    class_section = request.form.get("class", "CLASS")
    hour = request.form.get("hour", "HOUR")
    report_type = request.form.get("report_type", "both")
    date_str = datetime.now().strftime("%Y-%m-%d")
    
    # Clear existing photos to ensure fresh processing
    if os.path.exists(GROUP_PHOTOS_DIR):
        shutil.rmtree(GROUP_PHOTOS_DIR)
    os.makedirs(GROUP_PHOTOS_DIR, exist_ok=True)

    # Save uploaded photos
    photos = request.files.getlist('photos')
    for photo in photos:
        if photo:
            photo.save(os.path.join(GROUP_PHOTOS_DIR, photo.filename))
    
    attendance = {}

    for img_name in os.listdir(GROUP_PHOTOS_DIR):
        img_path = os.path.join(GROUP_PHOTOS_DIR, img_name)
        img = cv2.imread(img_path)
        if img is None:
            continue

        faces = model.get(img)
        for face in faces:
            emb = face.embedding / norm(face.embedding)
            sims = [cosine_sim(emb, k) for k in encodings]
            idx = int(np.argmax(sims))

            if sims[idx] >= ARC_THRESHOLD:
                roll = rolls[idx]
                if roll not in attendance:
                    attendance[roll] = {
                        "name": names[idx],
                        "time": datetime.now().strftime("%H:%M:%S")
                    }

    # -------- Absent Logic --------
    query = {}
    if "-" in class_section:
        import re
        match = re.search(r'\((.*?)\)', class_section)
        sec_code = match.group(1) if match else class_section
        dept, sec = sec_code.split("-", 1)
        query = {"department": dept, "section": sec}
    else:
        query = {"section": class_section}
    
    section_students = list(students_col.find(query))
    if not section_students and "(" in class_section:
        import re
        match = re.search(r'\((.*?)\)', class_section)
        sec_code = match.group(1) if match else class_section
        section_students = list(students_col.find({"section": sec_code}))

    section_rolls = {str(s["rollNo"]) for s in section_students}
    absent_rolls = sorted(section_rolls - set(attendance.keys()))

    # -------- Excel --------
    # Sanitize inputs for filename
    import re
    safe_course = re.sub(r'[^a-zA-Z0-9.\-]', '_', course)
    safe_hour = re.sub(r'[^a-zA-Z0-9.\-]', '_', hour)
    safe_class = re.sub(r'[^a-zA-Z0-9.\-]', '_', class_section)
    
    filename = f"{safe_course}_{safe_hour}_{safe_class}_{date_str}.xlsx"

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if report_type in ("present", "both"):
        ws_p = wb.create_sheet("Present")
        ws_p.append(["Roll No", "Name", "Time"])
        for r in sorted(attendance):
            ws_p.append([
                r, 
                attendance[r]["name"], 
                attendance[r]["time"]
            ])

    if report_type in ("absent", "both"):
        ws_a = wb.create_sheet("Absent")
        ws_a.append(["Roll No"])
        for r in absent_rolls:
            ws_a.append([r])

    wb.save(filename)

    # -------- Store in GridFS --------
    try:
        with open(filename, "rb") as f:
            file_id = fs.put(
                f,
                filename=filename,
                metadata={
                }
            )
        print(f"Group attendance Excel saved to GridFS with ID: {file_id}, filename: {filename}")

        # -------- Store in MongoDB Collection (for Dashboard/HOD) --------
        attendance_records_db = []
        
        # 1. Add Present Students
        for roll, data in attendance.items():
            record = {
                "rollNo": roll,
                "name": data["name"],
                "date": date_str,
                "subject": course,
                "className": class_section,
                "section": class_section.split("-")[-1].replace(")", "") if "-" in class_section else class_section,
                "status": "present",
                "time": data["time"],
                "type": "group",
                "facultyId": teacher.get("facultyId"),
                "facultyName": teacher.get("name"),
                "timestamp": datetime.utcnow()
            }
            attendance_records_db.append(record)

        # 2. Add Absent Students
        for roll in absent_rolls:
            student_doc = students_col.find_one({"rollNo": roll})
            student_name = student_doc["name"] if student_doc else "Unknown"
            
            record = {
                "rollNo": roll,
                "name": student_name,
                "date": date_str,
                "subject": course,
                "className": class_section,
                "section": class_section.split("-")[-1].replace(")", "") if "-" in class_section else class_section,
                "status": "absent",
                "time": None,
                "type": "group",
                "facultyId": teacher.get("facultyId"),
                "facultyName": teacher.get("name"),
                "timestamp": datetime.utcnow()
            }
            attendance_records_db.append(record)

        if attendance_records_db:
            # Upsert records
            for record in attendance_records_db:
                db["attendance"].update_one(
                    {
                        "rollNo": record["rollNo"],
                        "date": record["date"],
                        "className": record["className"]
                    },
                    {"$set": record},
                    upsert=True
                )
            print(f"Saved {len(attendance_records_db)} records to AttendanceDB.attendance")
    except Exception as e:
        print(f"Error saving Excel to GridFS: {str(e)}")
        return jsonify({"error": f"Failed to save report: {str(e)}"}), 500

    return jsonify({
        "message": "Group attendance completed",
        "filename": filename,
        "present": len(attendance),
        "absent": len(absent_rolls),
        "present_students": [{"roll": r, "name": attendance[r]["name"], "time": attendance[r]["time"]} for r in sorted(attendance)]
    }), 200
